#!/usr/bin/env python
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

from PIL import Image

def rgb_to_hex(r, g, b):
    """Converts RGB color values to a hex string."""
    return "{:02x}{:02x}{:02x}".format(r, g, b)

if __name__ == '__main__':
    # Load Image
    img = Image.open('rick-roll.jpg')
    pixels = img.load
    # Create a new workbook
    wb = Workbook()

    # Get the active worksheet (the default sheet)
    ws = wb.active

    image_width, image_height = img.size
    number_of_image_pixels_to_xcel_cells = 5
    image_width = int(image_width/number_of_image_pixels_to_xcel_cells)
    image_height = int(image_height / number_of_image_pixels_to_xcel_cells)
    # Set width of each cell
    for col in range(1,image_width+1):
        ws.column_dimensions[get_column_letter(col)].width = 2
    for rows in range(1,image_height + 1):
        for cols in range(1, image_width + 1):
            current_pixel_location = ((cols - 1) * number_of_image_pixels_to_xcel_cells,(rows -1) * number_of_image_pixels_to_xcel_cells)
            fill_color = rgb_to_hex(*img.getpixel(current_pixel_location))
            fill = PatternFill(fgColor=fill_color, fill_type="solid")
            ws.cell(row=rows, column=cols).value = ' ' # Have to put something in the cell or the fill does not work
            ws.cell(row=rows, column=cols).fill = fill

    # Rename the sheet
    ws.title = "Never Gonna Give You Up"

    # Save the workbook
    wb.save("never-going-give-you-up.xlsx")